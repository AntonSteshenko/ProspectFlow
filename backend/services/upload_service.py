"""
File upload service for handling CSV and XLSX files.

Provides validation, preview, and header extraction functionality.
"""
import csv
import io
from typing import List, Dict
import openpyxl


class UploadService:
    """
    Service for file upload operations.

    Methods:
        validate_file: Validate file format and size
        parse_preview: Extract first 5 rows for preview
        get_column_headers: Extract column headers from file
    """

    ALLOWED_EXTENSIONS = ['.csv', '.xlsx', '.xls']
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

    @classmethod
    def validate_file(cls, file) -> tuple[bool, str]:
        """
        Validate uploaded file.

        Args:
            file: Uploaded file object

        Returns:
            tuple: (is_valid, error_message)
        """
        # Check file extension
        ext = file.name.lower().split('.')[-1]
        if f'.{ext}' not in cls.ALLOWED_EXTENSIONS:
            return False, f"Unsupported file type. Allowed: {', '.join(cls.ALLOWED_EXTENSIONS)}"

        # Check file size
        if file.size > cls.MAX_FILE_SIZE:
            size_mb = file.size / (1024 * 1024)
            return False, f"File too large ({size_mb:.2f}MB). Maximum is 10MB."

        return True, ""

    @classmethod
    def get_column_headers(cls, file) -> List[str]:
        """
        Extract column headers from CSV or XLSX file.

        Args:
            file: Uploaded file object

        Returns:
            list: Column header names

        Raises:
            ValueError: If file format is unsupported or corrupted
        """
        ext = file.name.lower().split('.')[-1]

        try:
            if ext == 'csv':
                return cls._get_csv_headers(file)
            elif ext in ['xlsx', 'xls']:
                return cls._get_xlsx_headers(file)
            else:
                raise ValueError(f"Unsupported file extension: {ext}")
        except Exception as e:
            raise ValueError(f"Error reading file headers: {str(e)}")

    @classmethod
    def parse_preview(cls, file, num_rows=5) -> Dict:
        """
        Parse first N rows for preview.

        Args:
            file: Uploaded file object
            num_rows: Number of rows to preview (default 5)

        Returns:
            dict: {
                'headers': List[str],
                'rows': List[Dict],
                'total_rows': int (approximate)
            }

        Raises:
            ValueError: If file format is unsupported or corrupted
        """
        ext = file.name.lower().split('.')[-1]

        try:
            if ext == 'csv':
                return cls._parse_csv_preview(file, num_rows)
            elif ext in ['xlsx', 'xls']:
                return cls._parse_xlsx_preview(file, num_rows)
            else:
                raise ValueError(f"Unsupported file extension: {ext}")
        except Exception as e:
            raise ValueError(f"Error parsing file: {str(e)}")

    @classmethod
    def _get_csv_headers(cls, file) -> List[str]:
        """Extract headers from CSV file."""
        file.seek(0)
        content = file.read().decode('utf-8-sig')  # Handle BOM
        file.seek(0)

        reader = csv.DictReader(io.StringIO(content))
        return reader.fieldnames or []

    @classmethod
    def _get_xlsx_headers(cls, file) -> List[str]:
        """Extract headers from XLSX file."""
        file.seek(0)
        workbook = openpyxl.load_workbook(file, read_only=True)
        sheet = workbook.active

        # Get first row as headers
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value is not None else '')

        workbook.close()
        file.seek(0)
        return headers

    @classmethod
    def _parse_csv_preview(cls, file, num_rows=5) -> Dict:
        """Parse CSV file preview."""
        file.seek(0)
        content = file.read().decode('utf-8-sig')  # Handle BOM
        file.seek(0)

        reader = csv.DictReader(io.StringIO(content))
        headers = reader.fieldnames or []

        rows = []
        for i, row in enumerate(reader):
            if i >= num_rows:
                break
            rows.append(row)

        # Estimate total rows (not exact for CSV)
        total_rows = len(content.split('\n')) - 1  # Subtract header

        return {
            'headers': headers,
            'rows': rows,
            'total_rows': total_rows
        }

    @classmethod
    def _parse_xlsx_preview(cls, file, num_rows=5) -> Dict:
        """Parse XLSX file preview."""
        file.seek(0)
        workbook = openpyxl.load_workbook(file, read_only=True)
        sheet = workbook.active

        # Get headers (first row)
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value is not None else '')

        # Get preview rows
        rows = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if i >= num_rows:
                break
            row_dict = {header: value for header, value in zip(headers, row)}
            rows.append(row_dict)

        total_rows = sheet.max_row - 1  # Subtract header row

        workbook.close()
        file.seek(0)

        return {
            'headers': headers,
            'rows': rows,
            'total_rows': total_rows
        }
